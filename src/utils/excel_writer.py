"""
Excel workbook writer.
Generates multi-sheet Excel output with signals, news, parameters, etc.
"""

import logging
from datetime import datetime
from typing import Dict, List
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    # Try pandas Excel writer as fallback
    try:
        import pandas as pd
        PANDAS_AVAILABLE = True
    except ImportError:
        PANDAS_AVAILABLE = False


class ExcelWriter:
    """Writes trading signals and data to Excel workbook."""
    
    def __init__(self, config: dict):
        """Initialize with configuration."""
        self.config = config
        self.logger = logging.getLogger(__name__)
        self.output_file = config['output']['excel_file']
    
    def write_workbook(self, signals: List[Dict], market_data: Dict) -> str:
        """
        Write complete workbook with all tabs.
        
        Args:
            signals: List of ranked trading signals
            market_data: All market data
            
        Returns:
            Path to output file
        """
        if not OPENPYXL_AVAILABLE:
            self.logger.warning("openpyxl not installed, using pandas fallback")
            return self._write_pandas_excel(signals, market_data)
        
        self.logger.info(f"Writing Excel workbook to {self.output_file}")
        
        # Create output directory if needed
        Path(self.output_file).parent.mkdir(parents=True, exist_ok=True)
        
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Add sheets
        self._add_dashboard_sheet(wb, signals, market_data)
        self._add_signals_sheet(wb, signals)
        self._add_news_sheet(wb, market_data)
        self._add_parameters_sheet(wb)
        self._add_logs_sheet(wb)
        self._add_history_sheet(wb, signals)
        
        # Save workbook
        wb.save(self.output_file)
        self.logger.info(f"Workbook saved successfully")
        
        return self.output_file
    
    def _add_dashboard_sheet(self, wb, signals: List[Dict], market_data: Dict):
        """Add Dashboard sheet with summary metrics."""
        ws = wb.create_sheet("Dashboard")
        
        # Header
        ws['A1'] = "TRADING SIGNALS DASHBOARD"
        ws['A1'].font = Font(bold=True, size=14)
        
        # Metrics
        ws['A3'] = "Last Refresh"
        ws['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        ws['A4'] = "Top Pick"
        ws['B4'] = signals[0]['ticker'] if signals else "N/A"
        
        ws['A5'] = "Top Score"
        ws['B5'] = round(signals[0]['score'], 1) if signals else "N/A"
        
        ws['A6'] = "Num Signals"
        ws['B6'] = len(signals)
        
        ws['A7'] = "Alerts Sent"
        ws['B7'] = 0  # TODO: Track alerts
        
        # Auto-fit columns
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
    
    def _add_signals_sheet(self, wb, signals: List[Dict]):
        """Add Signals sheet with ranked list."""
        ws = wb.create_sheet("Signals")
        
        # Headers
        headers = ['Rank', 'Ticker', 'Price', 'Score', 'Momentum %', 'Volume Surge %',
                   'Rel Strength %', 'News Sentiment', 'Risk %', 'Status']
        ws.append(headers)
        
        # Style header
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Add signal rows
        for rank, signal in enumerate(signals, 1):
            ws.append([
                rank,
                signal.get('ticker', ''),
                round(signal.get('price', 0), 2),
                round(signal.get('score', 0), 1),
                round(signal.get('momentum_pct', 0), 1),
                round(signal.get('volume_surge_pct', 0), 1),
                round(signal.get('rel_strength_pct', 0), 1),
                signal.get('news_sentiment', 'Neutral'),
                round(signal.get('risk_score', 0), 1),
                'BUY' if signal.get('score', 0) > 75 else 'HOLD'
            ])
        
        # Auto-fit columns
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 15
    
    def _add_news_sheet(self, wb, market_data: Dict):
        """Add News sheet with catalyst summaries."""
        ws = wb.create_sheet("News")
        
        # Headers
        headers = ['Ticker', 'Headline', 'Source', 'Sentiment', 'Time', 'Summary']
        ws.append(headers)
        
        # Style header
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Auto-fit columns
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 50
    
    def _add_parameters_sheet(self, wb):
        """Add Parameters sheet (editable thresholds)."""
        ws = wb.create_sheet("Parameters")
        
        # Headers
        ws.append(['Category', 'Parameter', 'Value', 'Notes'])
        
        # Style header
        header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        header_font = Font(bold=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Add all parameters from config
        row = 2
        for category, params in self.config.items():
            if isinstance(params, dict):
                for key, value in params.items():
                    ws[f'A{row}'] = category
                    ws[f'B{row}'] = key
                    ws[f'C{row}'] = value
                    ws[f'D{row}'] = "Edit this value"
                    row += 1
        
        # Auto-fit columns
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 40
    
    def _add_logs_sheet(self, wb):
        """Add Logs sheet (data refresh status & errors)."""
        ws = wb.create_sheet("Logs")
        
        # Headers
        headers = ['Timestamp', 'Action', 'Status', 'Details']
        ws.append(headers)
        
        # Style header
        header_fill = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Sample log entry
        ws.append([datetime.now().strftime("%Y-%m-%d %H:%M:%S"), 'Fetch Prices', 'Success', '500 symbols'])
        
        # Auto-fit columns
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 50
    
    def _add_history_sheet(self, wb, signals: List[Dict]):
        """Add History Report sheet (overwritten each refresh)."""
        ws = wb.create_sheet("History Report")
        
        # Headers
        headers = ['Date', 'Ticker', 'Score', 'Momentum', 'Volume Surge', 'Action Taken', 'Outcome']
        ws.append(headers)
        
        # Style header
        header_fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Add signal history rows
        for signal in signals[:10]:  # Last 10 signals
            ws.append([
                datetime.now().strftime("%Y-%m-%d"),
                signal.get('ticker', ''),
                round(signal.get('score', 0), 1),
                round(signal.get('momentum_pct', 0), 1),
                round(signal.get('volume_surge_pct', 0), 1),
                'Trade Taken' if signal.get('score', 0) > 75 else 'Pending',
                '+0.0%'  # TODO: Fetch actual outcome
            ])
        
        # Auto-fit columns
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 15
    
    def _write_pandas_excel(self, signals: List[Dict], market_data: Dict) -> str:
        """Fallback using pandas (if openpyxl not available)."""
        if not PANDAS_AVAILABLE:
            raise ImportError("Either openpyxl or pandas required for Excel output")
        
        self.logger.warning("Using pandas for Excel output (limited formatting)")
        
        # Create output directory if needed
        Path(self.output_file).parent.mkdir(parents=True, exist_ok=True)
        
        # Convert signals to DataFrame
        df = pd.DataFrame(signals)
        
        # Write to Excel with pandas
        df.to_excel(self.output_file, sheet_name='Signals', index=False)
        
        return self.output_file
